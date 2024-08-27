import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

df1 = pd.read_excel('.Data/raw_data_separate_source/raw_data_chinese_wechat.xlsx')
df2 = pd.read_excel('.Data/raw_data_separate_source/raw_data_chinese_telephone.xlsx')

df1['序号'] = range(1, len(df1) + 1)
df2['序号'] = range(len(df1) + 1, len(df1) + len(df2) + 1)

df_combined = pd.concat([df1, df2], ignore_index=True)

wb = Workbook()
ws = wb.active
ws.title = "Merged Data"

ws.append(df_combined.columns.tolist())

for r in df_combined.itertuples(index=False, name=None):
    ws.append(r)

font_times = Font(name="宋体")
for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font_times
        cell.alignment = Alignment(horizontal='left')

save_path = "../Data/raw_data_total/raw_data_chinese_total.xlsx"
wb.save(save_path)
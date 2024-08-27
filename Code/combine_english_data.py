import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

df1 = pd.read_excel('.Data/raw_data_separate_source/raw_data_english_wechat.xlsx')
df2 = pd.read_excel('.Data/raw_data_separate_source/raw_data_english_telephone.xlsx')

df1['No.'] = range(1, len(df1) + 1)
df2['No.'] = range(len(df1) + 1, len(df1) + len(df2) + 1)

df_combined = pd.concat([df1, df2], ignore_index=True)

wb = Workbook()
ws = wb.active
ws.title = "Merged Data"

ws.append(df_combined.columns.tolist())

for r in df_combined.itertuples(index=False, name=None):
    ws.append(r)

font_times = Font(name="Times New Roman")
for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        cell.font = font_times
        cell.alignment = Alignment(horizontal='left')

save_path = "../Data/raw_data_total/raw_data_english_total.xlsx"
wb.save(save_path)
